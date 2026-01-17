import os
import json
import csv
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from routes.auth import get_current_user
from models.database import get_db_connection

export_bp = Blueprint('export', __name__)


def get_user_file(user_id: int, file_id: int):
    """Get a file from the database if it belongs to the user."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT * FROM generated_files WHERE id = ? AND user_id = ?
    ''', (file_id, user_id))
    
    row = cursor.fetchone()
    conn.close()
    
    return row


@export_bp.route('/api/download/excel/<int:file_id>', methods=['GET'])
def download_excel(file_id):
    """Download test cases as Excel file."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    row = get_user_file(user.id, file_id)
    if not row:
        return jsonify({'error': 'File not found'}), 404
    
    try:
        test_cases_data = json.loads(row['test_cases'])
        test_cases = test_cases_data.get('test_cases', [])
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Test ID", "Module", "Test Scenario", "Preconditions", 
            "Steps", "Test Data", "Expected Result", "Actual Result",
            "Status", "Priority", "Severity", "Edge Cases"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_num, tc in enumerate(test_cases, 2):
            data = [
                tc.get('test_id', ''),
                tc.get('module', ''),
                tc.get('test_scenario', ''),
                tc.get('preconditions', ''),
                tc.get('steps', ''),
                tc.get('test_data', ''),
                tc.get('expected_result', ''),
                tc.get('actual_result', ''),
                tc.get('status', 'Pending'),
                tc.get('priority', ''),
                tc.get('severity', ''),
                tc.get('edge_cases', '')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # Adjust column widths
        column_widths = [12, 20, 35, 25, 40, 20, 35, 20, 12, 12, 12, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{row['filename']}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        return jsonify({'error': 'Error generating Excel file'}), 500


@export_bp.route('/api/download/csv/<int:file_id>', methods=['GET'])
def download_csv(file_id):
    """Download test cases as CSV file."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    row = get_user_file(user.id, file_id)
    if not row:
        return jsonify({'error': 'File not found'}), 404
    
    try:
        test_cases_data = json.loads(row['test_cases'])
        test_cases = test_cases_data.get('test_cases', [])
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "Test ID", "Module", "Test Scenario", "Preconditions", 
            "Steps", "Test Data", "Expected Result", "Actual Result",
            "Status", "Priority", "Severity", "Edge Cases"
        ]
        writer.writerow(headers)
        
        # Data rows
        for tc in test_cases:
            writer.writerow([
                tc.get('test_id', ''),
                tc.get('module', ''),
                tc.get('test_scenario', ''),
                tc.get('preconditions', ''),
                tc.get('steps', '').replace('\n', ' | '),
                tc.get('test_data', ''),
                tc.get('expected_result', ''),
                tc.get('actual_result', ''),
                tc.get('status', 'Pending'),
                tc.get('priority', ''),
                tc.get('severity', ''),
                tc.get('edge_cases', '')
            ])
        
        # Get CSV content
        csv_content = output.getvalue()
        output.close()
        
        filename = f"{row['filename']}.csv"
        
        return Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
    except Exception as e:
        print(f"CSV export error: {str(e)}")
        return jsonify({'error': 'Error generating CSV file'}), 500


@export_bp.route('/api/history', methods=['GET'])
def get_history():
    """Get user's generated file history."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT id, filename, project_type, created_at 
        FROM generated_files 
        WHERE user_id = ? 
        ORDER BY created_at DESC
        LIMIT 50
    ''', (user.id,))
    
    rows = cursor.fetchall()
    conn.close()
    
    history = []
    for row in rows:
        history.append({
            'id': row['id'],
            'filename': row['filename'],
            'project_type': row['project_type'],
            'created_at': row['created_at']
        })
    
    return jsonify({'history': history}), 200


@export_bp.route('/api/history/<int:file_id>', methods=['DELETE'])
def delete_history_item(file_id):
    """Delete a history item."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        DELETE FROM generated_files WHERE id = ? AND user_id = ?
    ''', (file_id, user.id))
    
    deleted = cursor.rowcount > 0
    conn.commit()
    conn.close()
    
    if deleted:
        return jsonify({'message': 'File deleted successfully'}), 200
    else:
        return jsonify({'error': 'File not found'}), 404
